import os
import argparse
import logging
import re
import time
import openpyxl
import datetime
import ipaddress
import validators
import json
from config import CommandEnumDef, CommandEnumAgg, CommandEnumCau
from concurrent.futures import ProcessPoolExecutor, as_completed

# logging.basicConfig(level=logging.INFO, filename="logs.log", filemode="w", format='%(asctime)s - %(levelname)s - %(message)s')
# auditoria per servidor (sistema de logs)
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# globals
logger = logging.getLogger()
SAVES = "results/temp/"
def welcome():
    intro_message = """
*******************************************************************
 _  __         _       
| |/ /   _  __| | ___  
| ' / | | |/ _` |/ _ \ 
| . \ |_| | (_| | (_) |
|_|\_\__,_|\__,_|\___/ 
                                       
 Kudo 1.0                                          
 Coded by OriJari                              
 https://github.com/OriJari                              
                                      
*******************************************************************
usage: kudo.py [-h] [-i IP] [-d DOMAIN] [-lI LIST_IP] [-lD LIST_DOMAIN] [-r] [-v] [-a] [-c]

Kudo is a script that automates the recognition and scanning of services, technologies and vulnerabilities.

*******************************************************************
    """
    print(intro_message)
def validate_ip(ip):
    try:
        ipaddress.ip_address(ip)
        return True
    except ValueError:
        return False

def validate_domain(domain):
    try:
        validators.domain(domain)
        return True
    except ValueError:
        return False

def expand_ip_range(cidr):
    try:
        network = ipaddress.ip_network(cidr, strict=False)
        return [str(ip) for ip in network.hosts()]
    except ValueError:
        return []

def treat_json(filename, data):
    with open(filename, 'w') as f:
        for item in data:
            if not item == "127.0.0.1":
                f.write("%s\n" % item)

def save_info(source, output, target):
    if source == "harvester":
        for section, content in output.items():
            if isinstance(content, list):
                treat_json(f'{SAVES}{target}_harvester_{section}.txt', content)

def merge_unique_ips(file1, file2, output_file):
    ips_set = set()
    with open(file1, 'r') as f1:
        for line in f1:
            ip = line.strip()
            if ip:
                ips_set.add(ip)

    with open(file2, 'r') as f2:
        for line in f2:
            ip = line.strip()
            if ip:
                ips_set.add(ip)

    with open(output_file, 'w') as output:
        for ip in sorted(ips_set):
            output.write(f"{ip}\n")

def merge_unique_subdomain(file_subfinder, file_gobuster, output_file):
    with open(file_subfinder, 'r') as f:
        subfinder_lines = f.read().splitlines()

    with open(file_gobuster, 'r') as f:
        gobuster_lines = [line.replace('Found: ', '').strip() for line in f]

    combined_lines = list(set(subfinder_lines + gobuster_lines))

    with open(output_file, 'w') as f:
        for line in sorted(combined_lines):
            f.write(line + '\n')

def execute_order_66(command):
    return os.popen(command).read()

def exec_dmitry(target):
    logger.info(f"[·] Dmitry for {target} started.")
    result_dmirty = execute_order_66(CommandEnumDef.DMIRTY.format(target,target))
    logger.info(result_dmirty)
    logger.info(f"[·] Dmitry for {target} ended.")

def exec_harvester(target, name_file_target):
    logger.info(f"[·] theHarvester for {target} started.")
    result_harvester = execute_order_66(CommandEnumDef.HARVESTER.format(target, name_file_target))
    logger.info(result_harvester)
    with open(f'{SAVES}{name_file_target}_harvester.json', 'r') as file:
        output = json.load(file)
    save_info("harvester", output, name_file_target)
    logger.info(f"[·] theHarvester for {target} ended.")

def exec_subfinder(target,flags, name_file_target):
    logger.info(f"[·] subfinder for {target} started.")
    if flags.cautious:
        result_subfinder = execute_order_66(CommandEnumCau.SUBFINDER.format(target, name_file_target))
    else:
        result_subfinder = execute_order_66(CommandEnumDef.SUBFINDER.format(target, name_file_target))
    logger.info(result_subfinder)
    logger.info(f"[·] subfinder for {target} ended.")

def exec_dnsx(target,flags):
    logger.info(f"[·] DNSX for {target} started.")
    if flags.aggresive and not flags.cautious:
        result_dnsx = execute_order_66(CommandEnumAgg.DNSX.format(f"{SAVES}{target}_total_subdomains{target}.txt", target))
        execute_order_66(CommandEnumAgg.DNSX2.format(f"{SAVES}{target}_total_subdomains{target}.txt", target))
    elif flags.cautious and not flags.aggressive:
        result_dnsx = execute_order_66(CommandEnumCau.DNSX.format(f"{SAVES}{target}_total_subdomains{target}.txt", target))
        execute_order_66(CommandEnumCau.DNSX2.format(f"{SAVES}{target}_total_subdomains{target}.txt", target))
    else:
        result_dnsx = execute_order_66(CommandEnumDef.DNSX.format(f"{SAVES}{target}_total_subdomains{target}.txt", target))
        execute_order_66(CommandEnumDef.DNSX2.format(f"{SAVES}{target}_total_subdomains{target}.txt", target))
    logger.info(result_dnsx)
    logger.info(f"[·] DNSX for {target} ended.")

def exec_nmap(target, flags):
    logger.info(f"[·] NMAP for {target} started.")
    if flags.domain or flags.list_Domain:
        if flags.aggressive and not flags.cautious:
            execute_order_66(CommandEnumAgg.NMAPLIST.format(f"{SAVES}{target}_total_ips.txt",target))
        elif flags.cautious and not flags.aggressive:
            execute_order_66(CommandEnumCau.NMAPLIST.format(f"{SAVES}{target}_total_ips.txt",target))
        else:
            execute_order_66(CommandEnumDef.NMAPLIST.format(f"{SAVES}{target}_total_ips.txt",target))
    elif flags.ip or flags.list_Ip:
        if flags.aggressive and not flags.cautious:
            execute_order_66(CommandEnumAgg.NMAP.format(target,target))
        elif flags.cautious and not flags.aggressive:
            execute_order_66(CommandEnumCau.NMAP.format(target,target))
        else:
            execute_order_66(CommandEnumDef.NMAP.format(target,target))
    with open(f"results/temp/{target}_nmap.txt", 'r') as file:
        content = file.read()
    logger.info(content)
    logger.info(f"[·] NMAP for {target} ended.")

def exec_ferox(target,name_file_target):
    logger.info(f"[·] Feroxbuster for {target} started.")
    result_ferox = execute_order_66(CommandEnumDef.FEROXBUSTER.format(target,name_file_target))
    logger.info(result_ferox)
    logger.info(f"[·] Feroxbuster for {target} ended.")

def exec_gobuster(target, name_file_target):
    logger.info(f"[·] Gobuster DNS for {target} starterd.")
    result_gobuster_dns = execute_order_66(CommandEnumDef.GOBUSTERDNS.format(target,name_file_target))
    logger.info(result_gobuster_dns)
    logger.info(f"[·] Gobuster DNS for {target} ended.")

def exec_waf(target, name_file_target):
    logger.info(f"[·] Waffw00f for {target} started.")
    result_waf = execute_order_66(CommandEnumDef.WAF.format(target,name_file_target))
    logger.info(result_waf)
    logger.info(f"[·] Waffw00f for {target} ended.")

def check_scan_aborted(file_path):
    with open(file_path, 'r') as file:
        data = file.readlines()
    return "scan_aborted" in data
def exec_wpscan(target, flags, name_file_target):
    logger.info(f"[·] WPscan Vuln for {target} started.")
    if flags.aggressive and not flags.cautious:
        content =execute_order_66(CommandEnumAgg.WPSACN.format(target,name_file_target))
        if check_scan_aborted(f"{SAVES}{target}_wpscan.txt"):
            content = execute_order_66(CommandEnumAgg.WPSACN.format(f"www.{target}", name_file_target))
    elif flags.cautious and not flags.aggressive:
        content =execute_order_66(CommandEnumCau.WPSACN.format(target,name_file_target))
        if check_scan_aborted(f"{SAVES}{target}_wpscan.txt"):
            content = execute_order_66(CommandEnumCau.WPSACN.format(f"www.{target}", name_file_target))
    else:
        content = execute_order_66(CommandEnumDef.WPSACN.format(target,name_file_target))
        if check_scan_aborted(f"{SAVES}{target}_wpscan.txt"):
            content = execute_order_66(CommandEnumDef.WPSACN.format(f"www.{target}", name_file_target))
    with open(f"results/temp/{target}_wpscan.txt", 'r') as file:
        content = file.read()
    logger.info(content)
    logger.info(f"[·] WPscan for {target} ended.")

def work_domini(targets, flags):
    for target in targets:
        if validate_domain(target):

            logger.info(f"[+] OSINT and Recon for {target} started.")
            exec_dmitry(target)
            exec_harvester(target, target)
            exec_subfinder(target,flags, target)
            exec_gobuster(target, target)

            merge_unique_subdomain(f"{SAVES}{target}_subfinder_subdomain.txt",f"{SAVES}{target}_gobuster_dns.txt", f'{SAVES}{target}_total_subdomains.txt')
            dnsx_break = False
            try:
                 exec_dnsx(target,flags)
            except:
                logger.error("[-] DNSX failed")
                dnsx_break = True

            if not dnsx_break:
                merge_unique_ips(f'{SAVES}{target}_harvester_ips.txt', f'{SAVES}{target}_dnsx2_subdomains.txt',f'{SAVES}{target}_total_ips.txt')
            else:
                os.system(f"cp {SAVES}{target}_harvester_ips.txt {SAVES}{target}_total_ips.txt")

            exec_nmap(target, flags)
            exec_waf(target, target)
            exec_wpscan(target, flags, target)

            logger.info(f"[+] OSINT and Recon for {target} completed.")
        else:
            logger.error(f"[-] Domain {target} not valid or not reachable")

def scan_recon_ip(target, flags):
    global domain_target
    time.sleep(0.5)
    exec_dmitry(target)
    exec_nmap(target, flags)


    domain_target = execute_order_66(f"echo {target} | dnsx -silent -resp-only -ptr ").strip()
    if domain_target == "":
        dnsx_worked = False
        logger.error(f"[-] DNSX couldn't resolve {target} ")
    else:
        dnsx_worked = True
        execute_order_66(f"echo {target} | dnsx -silent -re -ptr -o {SAVES}{target}_dnsx_resolve_domain").strip()

    if dnsx_worked:
        if validate_domain(domain_target):
            exec_harvester(domain_target,target)
            exec_subfinder(domain_target,flags,target)
            exec_gobuster(domain_target, target)

            #just to see what finds, for the rest of the code of recon ip is not critical
            merge_unique_subdomain(f"{SAVES}{target}_subfinder_subdomain.txt", f"{SAVES}{target}_gobuster_dns.txt",
                                   f'{SAVES}{target}_total_subdomains.txt')
            dnsx_break = False
            try:
                exec_dnsx(target, flags)
            except:
                logger.error("[-] DNSX failed")
                dnsx_break = True

            if not dnsx_break:
                merge_unique_ips(f'{SAVES}{target}_harvester_ips.txt', f'{SAVES}{target}_dnsx2_subdomains.txt',
                                 f'{SAVES}{target}_total_ips.txt')
            else:
                os.system(f"cp {SAVES}{target}_harvester_ips.txt {SAVES}{target}_total_ips.txt")

            exec_waf(domain_target, target)
            exec_wpscan(domain_target, flags, target)


    else:
        exec_gobuster(target,target)

def work_ips(targets, flags):
    for target in targets:
        if '/' in target:
            ip_range = expand_ip_range(target)
            for ip in ip_range:
                if validate_ip(ip):
                    logger.info(f"[+] OSINT and Recon for {ip} started.")
                    scan_recon_ip(ip, flags)
                    logger.info(f"[+] OSINT and Recon for {ip} completed.")

                else:
                    logger.error(f"[-] IP {ip} not valid or not reachable")
        else:
            if validate_ip(target):
                logger.info(f"[+] OSINT and Recon for {target} started.")
                scan_recon_ip(target, flags)
                logger.info(f"[+] OSINT and Recon for {target} completed.")

            else:
                logger.error(f"[-] IP {target} not valid or not reachable")

def recon(flags):
    if flags.domain:
        targets = [flags.domain]
        work_domini(targets, flags)
    elif flags.list_Domain:
        try:
            with open(flags.list_Domain, 'r') as file:
                targets = [line.strip() for line in file.readlines() if line.strip()]
            work_domini(targets, flags)
        except FileNotFoundError:
            logger.error(f"[-] File not found: {flags.list_Ip or flags.list_Domain}")
            return
    elif flags.ip:
        targets = [flags.ip]
        work_ips(targets, flags)
    elif flags.list_Ip:
        try:
            with open(flags.list_Ip, 'r') as file:
                targets = [line.strip() for line in file.readlines() if line.strip()]
            work_ips(targets, flags)
        except FileNotFoundError:
            logger.error(f"[-] File not found: {flags.list_Ip or flags.list_Domain}")
            return
    else:
        logger.error("[-] No target specified. Use -i , -d , -lI or -lD to specify the IP(s) or the Domain(s).")
        return

def exec_nuclei(target,flags):
    logger.info(f"[·] Nuclei for {target} starterd.")
    if not flags.vuln_scan and not flags.recon: # no flags, default
        result_nuclei = execute_order_66(CommandEnumDef.NUCLEI.format(f"{SAVES}{target}_total_subdomains.txt", target))
    elif not flags.recon: # -vuln_scan, only
        result_nuclei = execute_order_66(CommandEnumDef.NUCLEI.format(target, target))
    else: # -recon i -vuln_scan, default
        result_nuclei = execute_order_66(CommandEnumDef.NUCLEI.format(f"{SAVES}{target}_total_subdomains.txt", target))

    logger.info(result_nuclei)
    logger.info(f"[·] Gobuster DNS for {target} ended.")

def exec_wpscan_vuln(target, flags, name_file_target):
    logger.info(f"[·] WPscan Vuln for {target} started.")
    if flags.aggressive and not flags.cautious:
        content = execute_order_66(CommandEnumAgg.WPSCANVULN.format(target,name_file_target))
        if check_scan_aborted(f"{SAVES}{target}_wpscanvuln.txt"):
            content = execute_order_66(CommandEnumAgg.WPSCANVULN.format(f"www.{target}", name_file_target))
    elif flags.cautious and not flags.aggressive:
        content = execute_order_66(CommandEnumCau.WPSCANVULN.format(target,name_file_target))
        if check_scan_aborted(f"{SAVES}{target}_wpscanvuln.txt"):
            content = execute_order_66(CommandEnumCau.WPSCANVULN.format(f"www.{target}", name_file_target))
    else:
        content = execute_order_66(CommandEnumDef.WPSCANVULN.format(target,name_file_target))
        if check_scan_aborted(f"{SAVES}{target}_wpscanvuln.txt"):
            content = execute_order_66(CommandEnumDef.WPSCANVULN.format(f"www.{target}", name_file_target))
    with open(f"results/temp/{target}_wpscanvuln.txt", 'r') as file:
        content = file.read()
    logger.info(content)
    logger.info(f"[·] WPscan Vuln for {target} ended.")

def vuln_domini(targets, flags):
    for target in targets:
        if validate_domain(target):

            logger.info(f"[+] Vulnerability sacn for {target} started.")

            exec_nuclei(target, flags)
            exec_wpscan_vuln(target,flags,target)


            logger.info(f"[+] Vulnerability sacn {target} completed.")
        else:
            logger.error(f"[-] Domain {target} not valid or not reachable")

def vuln_scan_ip(target,flags):
        logger.info(f"[+] Vulnerability sacn for {target} started.")

        exec_nuclei(target, flags)
        domain_target = execute_order_66(f"echo {target} | dnsx -silent -resp-only -ptr ").strip()
        if domain_target == "":
            dnsx_worked = False
            logger.error(f"[-] DNSX couldn't resolve {target} ")
        else:
            dnsx_worked = True
            execute_order_66(
                f"echo {target} | dnsx -silent -re -ptr -o {SAVES}{target}_dnsx_resolve_domain").strip()

        if dnsx_worked:
            exec_wpscan_vuln(target, flags, target)

        logger.info(f"[+] Vulnerability sacn {target} completed.")

def vuln_ip(targets, flags):
    for target in targets:
        if '/' in target:
            ip_range = expand_ip_range(target)
            for ip in ip_range:
                if validate_ip(ip):
                    logger.info(f"[+] OSINT and Recon for {ip} started.")
                    vuln_scan_ip(ip, flags)
                    logger.info(f"[+] OSINT and Recon for {ip} completed.")

                else:
                    logger.error(f"[-] IP {ip} not valid or not reachable")
        else:
            if validate_ip(target):
                logger.info(f"[+] OSINT and Recon for {target} started.")
                vuln_scan_ip(target, flags)
                logger.info(f"[+] OSINT and Recon for {target} completed.")

            else:
                logger.error(f"[-] IP {target} not valid or not reachable")

def vuln(flags):
    if flags.domain:
        targets = [flags.domain]
        vuln_domini(targets, flags)
    elif flags.list_Domain:
        try:
            with open(flags.list_Domain, 'r') as file:
                targets = [line.strip() for line in file.readlines() if line.strip()]
            vuln_domini(targets, flags)
        except FileNotFoundError:
            logger.error(f"[-] File not found: {flags.list_Ip or flags.list_Domain}")
            return
    elif flags.ip:
        targets = [flags.ip]
        vuln_ip(targets, flags)
    elif flags.list_Ip:
        try:
            with open(flags.list_Ip, 'r') as file:
                targets = [line.strip() for line in file.readlines() if line.strip()]
            vuln_ip(targets, flags)
        except FileNotFoundError:
            logger.error(f"[-] File not found: {flags.list_Ip or flags.list_Domain}")
            return
    else:
        logger.error("[-] No target specified. Use -i , -d , -lI or -lD to specify the IP(s) or the Domain(s).")
        return

def clean_txt(file_path):
    with open(f"{SAVES}{file_path}", 'r', encoding='utf-8') as f:
        lines = f.readlines()
    return [line.strip() for line in lines]

def remove_illegal_characters(value):
    illegal_characters_re = re.compile(r'[\x00-\x08\x0B\x0C\x0E-\x1F]')

    if isinstance(value, str):
        return illegal_characters_re.sub('', value)
    return value
def process_files_sheet_txt(results_workbook, file_path):

    base_name = os.path.basename(file_path)
    parts = base_name.split('_')
    tool_name = parts[1].split('.')[0]


    ws_single_model = results_workbook.create_sheet(title=tool_name)

    set_columns_width(ws_single_model)

    data = clean_txt(file_path)
    for row in data:
        cleaned_row = [remove_illegal_characters(row)]
        try:
            ws_single_model.append(cleaned_row)
        except openpyxl.utils.exceptions.IllegalCharacterError as e:
            logger.info(f"Error: {e} en la línea: {cleaned_row}")

def clean_json(file_path):
    with open(file_path, 'r') as f:
        content = f.read().strip()
        if not content:
            return None
        data = json.loads(content)
    return data

def process_files_sheet_json(results_workbook, file_path):
    base_name = os.path.basename(file_path)
    parts = base_name.split('_')
    tool_name = parts[1].split('.')[0]

    ws_single_model = results_workbook.create_sheet(title=tool_name)
    set_columns_width(ws_single_model)

    data = clean_json(file_path)

    if isinstance(data, dict):
        list_values = [v for v in data.values() if isinstance(v, list)]
        if list_values:
            max_length = max(len(v) for v in list_values)
        else:
            max_length = 0
        headers = list(data.keys())
        ws_single_model.append(headers)

        for i in range(max_length):
            row = []
            for key in headers:
                if isinstance(data[key], list) and i < len(data[key]):
                    value = data[key][i]
                else:
                    value = None

                # Convertir valores a cadenas de texto si no son compatibles
                if not isinstance(value, (str, int, float, bool, type(None))):
                    value = str(value)

                row.append(value)

            ws_single_model.append(row)
def process_files_sheet_interesting(results_workbook, file_path):
    ws = results_workbook["Interesting Findings"]
    base_name = os.path.basename(file_path)
    parts = base_name.split('_')
    column_name = parts[2].split('.')[0]

    if column_name == "ips" and "total" not in base_name:
        return

    data = clean_txt(file_path)
    if not data:
        return

    col_idx = 1
    while ws.cell(row=1, column=col_idx).value is not None:
        col_idx += 1

    ws.cell(row=1, column=col_idx).value = column_name
    for row_idx, line in enumerate(data, start=2):
        ws.cell(row=row_idx, column=col_idx).value = line
def make_excel(targets, is_ip):
    for target in targets:
        results_workbook = openpyxl.Workbook()
        ws = results_workbook.active
        ws.title = "Interesting Findings"

        if is_ip:
            if '/' in target:
                ip_range = expand_ip_range(target)
                for ip in ip_range:
                    logger.info(f"[+] Generating Excel for {ip}.")
                    for root, dirs, files in os.walk(SAVES):
                        for file in files:
                            #txt
                            if file.startswith(ip) and file.endswith(
                                    ".txt") and "harvester" not in file and "dnsx2" not in file and "total" not in file:
                                file_path = os.path.join(file)
                                process_files_sheet_txt(results_workbook, file_path)
                            #json
                            elif file.startswith(ip) and file.endswith(".json"):
                                file_path = os.path.join(root, file)
                                process_files_sheet_json(results_workbook, file_path)

                            elif file.startswith(ip) and file.endswith(".txt") and ("harvester" in file or "total" in file):
                                file_path = os.path.join(file)
                                process_files_sheet_interesting(results_workbook, file_path)
                    output_file = f"results/kudo_{ip}_{datetime.datetime.now()}.xlsx"
                    results_workbook.save(output_file)
                    logger.info(f"[+] Excel for {ip} finished.")
            else:
                logger.info(f"[+] Generating Excel for {target}.")
                for root, dirs, files in os.walk(SAVES):
                    for file in files:
                        #txt
                        if file.startswith(target) and file.endswith(
                                ".txt") and "harvester" not in file and "dnsx2" not in file and "total" not in file:
                            file_path = os.path.join(file)
                            process_files_sheet_txt(results_workbook, file_path)
                        #json
                        elif file.startswith(target) and file.endswith(".json"):
                            file_path = os.path.join(root, file)
                            process_files_sheet_json(results_workbook, file_path)

                        elif file.startswith(target) and file.endswith(".txt") and ("harvester" in file or "total" in file):
                            file_path = os.path.join(file)
                            process_files_sheet_interesting(results_workbook, file_path)
                output_file = f"results/kudo_{target}_{datetime.datetime.now()}.xlsx"
                results_workbook.save(output_file)
                logger.info(f"[+] Excel for {target} finished.")


        else:
            logger.info(f"[+] Generating Excel for {target}.")
            for root, dirs, files in os.walk(SAVES):
                for file in files:
                    if file.startswith(target) and file.endswith(".txt") and "harvester" not in file and "dnsx2" not in file and "total" not in file :
                        file_path = os.path.join(file)
                        process_files_sheet_txt(results_workbook, file_path)
                    elif file.startswith(target) and file.endswith(".json"):
                        file_path = os.path.join(root, file)
                        process_files_sheet_json(results_workbook, file_path)
                    elif file.startswith(target) and file.endswith(".txt") and ("harvester" in file or "total" in file):
                        file_path = os.path.join(file)
                        process_files_sheet_interesting(results_workbook, file_path)

            output_file = f"results/kudo_{target}_{datetime.datetime.now()}.xlsx"
            results_workbook.save(output_file)
            logger.info(f"[+] Excel for {target} finished.")

def choose_excel(flags):
    if flags.domain:
        targets = [flags.domain]
        make_excel(targets, False)
    elif flags.list_Domain:
        with open(flags.list_Domain, 'r') as file:
            targets = [line.strip() for line in file.readlines() if line.strip()]
            make_excel(targets, False)

    elif flags.ip:
        targets = [flags.ip]
        make_excel(targets,  True)
    elif flags.list_Ip:
        with open(flags.list_Ip, 'r') as file:
            targets = [line.strip() for line in file.readlines() if line.strip()]
            make_excel(targets,  True)

def set_columns_width(ws):
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

def main(flags):
    os.popen("rm -rf results/temp/*")
    welcome()
    if not flags.vuln_scan and not flags.recon: # no flags, default
        recon(flags)
        vuln(flags)
    elif not flags.recon: # -vuln_scan, only
        vuln(flags)
    elif not flags.vuln_scan: # -recon, only
        recon(flags)
    else: # -recon i -vuln_scan, default
        recon(flags)
        vuln(flags)

    choose_excel(flags)

    os.popen("rm -rf results/temp/*")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Recon & Scan Vulns automated script tool")
    parser.add_argument("-i", "--ip", type=str, help="Target IP to scan")
    parser.add_argument("-d", "--domain", type=str, help="Target domain to scan")
    parser.add_argument("-lI", "--list_Ip", type=str, help="List of IPs to scan")
    parser.add_argument("-lD", "--list_Domain", type=str, help="List of domains to scan")
    parser.add_argument("-r", "--recon", action="store_true", help="Perform only reconnaissance")
    parser.add_argument("-v", "--vuln_scan", action="store_true", help="Perform only vulnerability scanning")
    parser.add_argument("-a", "--aggressive", action="store_true", help="Run scans in aggressive mode")
    parser.add_argument("-c", "--cautious", action="store_true", help="Run scans in cautious mode")

    args = parser.parse_args()
    main(args)
